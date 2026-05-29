"""Value cleaners + workbook loaders shared across the 4 file parsers."""
from __future__ import annotations

import re
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def load_workbook(source: bytes | Path | str) -> Workbook:
    """Load an .xlsx from bytes or a path."""
    if isinstance(source, (bytes, bytearray)):
        return openpyxl.load_workbook(BytesIO(source), data_only=True, read_only=False)
    return openpyxl.load_workbook(str(source), data_only=True, read_only=False)


def pick_first_sheet(wb: Workbook) -> Worksheet:
    """Return the first non-empty sheet, or the active sheet if all are empty."""
    for name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row and ws.max_row > 1:
            return ws
    return wb.active  # type: ignore[return-value]


def find_sheet_by_keyword(wb: Workbook, *keywords: str) -> Worksheet | None:
    """Find a sheet whose name contains any of the given keywords (case-insensitive)."""
    kws = [k.lower() for k in keywords]
    for name in wb.sheetnames:
        low = name.lower()
        if any(k in low for k in kws):
            return wb[name]
    return None


# ─────────────────────────────────────────────────────────────────────────────
#  Value cleaners
# ─────────────────────────────────────────────────────────────────────────────
_NA_STRINGS = {"#n/a", "n/a", "na", "#na", "#ref!", "#value!", "#name?"}


def is_na(v: Any) -> bool:
    """True if the cell holds an Excel error value or blank."""
    if v is None:
        return True
    if isinstance(v, str) and v.strip().lower() in _NA_STRINGS:
        return True
    return False


def clean_str(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str) and v.strip().lower() in _NA_STRINGS:
        return None
    s = str(v).strip()
    return s if s else None


def clean_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, str) and v.strip().lower() in _NA_STRINGS:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def clean_int(v: Any) -> int | None:
    if v is None or v == "":
        return None
    if isinstance(v, str) and v.strip().lower() in _NA_STRINGS:
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def clean_date(v: Any) -> str | None:
    """Returns an ISO date string (YYYY-MM-DD) or None."""
    if v is None or v == "":
        return None
    if hasattr(v, "isoformat"):
        return v.isoformat()[:10]
    s = str(v).strip()
    return s if s else None


def normalize_header(v: Any) -> str:
    """Lower-case, strip, collapse internal whitespace."""
    s = str(v or "").strip().lower()
    return re.sub(r"\s+", " ", s)
