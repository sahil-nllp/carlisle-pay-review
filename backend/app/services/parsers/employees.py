"""Parser for Employee_Details.xlsx — the master employee list.

Source columns (FY26 export):
  A  Employee Number          → emp_num
  B  Last Name                → last_name
  C  First Name               → first_name
  D  Preferred Name           → preferred_name
  E  Employee Category        → category    (e.g. "Part Time", "Casual")
  F  Work Email               → email
  G  Age / Date of Birth       → age (direct) or derived from dob
  H  Date Service Start       → service_start_date
  I  Date Hired               → hire_date
  J  Payroll Name Selection   → site
  K  Job Classification       → job_classification
  L  Rate Type                → rate_type   (e.g. "Hourly")
  M  Amount                   → current_rate
  N  Hours Per Pay Period     → hours_per_pay_period
  O  Award Agreement          → current_award

Tolerates re-ordered columns, optional columns, and blank rows.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date as date_type, datetime
from pathlib import Path
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.services.parsers._helpers import (
    clean_date,
    clean_float,
    clean_int,
    clean_str,
    load_workbook,
    normalize_header,
    pick_first_sheet,
)


# ─────────────────────────────────────────────────────────────────────────────
#  Header detection
# ─────────────────────────────────────────────────────────────────────────────
COLUMN_KEYWORDS: dict[str, list[str]] = {
    "emp_num":              ["employee number", "emp no", "emp #", "emp number"],
    "last_name":            ["last name", "surname"],
    "first_name":           ["first name"],
    "preferred_name":       ["preferred name", "known as", "preferred"],
    "category":             ["employee category", "category"],
    "email":                ["work email", "email"],
    "age":                  ["age"],
    "dob":                  ["date of birth", "dob", "birth date"],
    "service_start_date":   ["date service start", "service start", "service commenced"],
    "hire_date":            ["date hired", "hire date", "date hire"],
    "site":                 ["payroll name selection", "site", "location"],
    "job_classification":   ["job classification", "classification", "position"],
    "rate_type":            ["rate type"],
    "current_rate":         ["amount", "rate", "hourly rate"],
    "hours_per_pay_period": ["hours per pay period", "hours per pay", "hpp"],
    "current_award":        ["award agreement", "award"],
}


# ─────────────────────────────────────────────────────────────────────────────
#  Result types
# ─────────────────────────────────────────────────────────────────────────────
@dataclass
class ParsedEmployee:
    """One row from Employee_Details.xlsx."""

    emp_num: str
    first_name: str
    last_name: str
    site: str

    preferred_name: str | None = None
    email: str | None = None
    dob: str | None = None                   # ISO date string
    age: int | None = None                   # derived from dob at effective_date
    service_start_date: str | None = None    # ISO date string
    hire_date: str | None = None             # ISO date string
    category: str | None = None
    job_classification: str | None = None
    rate_type: str | None = None
    current_rate: float | None = None
    hours_per_pay_period: float | None = None
    hours_per_week: float | None = None      # derived from hours_per_pay_period / 2
    current_award: str | None = None

    is_departed: bool = False


@dataclass
class EmployeeParseResult:
    sheet_name: str
    column_map: dict[str, int]
    employees: list[ParsedEmployee] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


# ─────────────────────────────────────────────────────────────────────────────
#  Public API
# ─────────────────────────────────────────────────────────────────────────────
def parse_employee_details(
    source: bytes | Path | str,
    *,
    effective_date: date_type | None = None,
) -> EmployeeParseResult:
    """Parse Employee_Details.xlsx.

    Parameters
    ----------
    source
        Path, str, or raw bytes of the .xlsx file.
    effective_date
        Used to calculate `age` from `dob`. Defaults to today.
    """
    wb = load_workbook(source)
    sheet = pick_first_sheet(wb)
    header_row = _find_header_row(sheet)
    column_map = _detect_columns(sheet, header_row)

    required = {"emp_num", "first_name", "last_name", "site", "current_award"}
    missing = required - column_map.keys()
    if missing:
        raise ValueError(
            f"Employee_Details: required columns not found: {sorted(missing)}. "
            f"Detected: {sorted(column_map.keys())}"
        )

    result = EmployeeParseResult(sheet_name=sheet.title, column_map=column_map)
    eff = effective_date or date_type.today()

    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        emp = _row_to_employee(row, column_map, eff)
        if emp is None:
            continue
        result.employees.append(emp)

    if not result.employees:
        result.warnings.append("No employee rows found.")

    return result


# ─────────────────────────────────────────────────────────────────────────────
#  Internals
# ─────────────────────────────────────────────────────────────────────────────
def _find_header_row(sheet: Worksheet) -> int:
    """Scan first 5 rows looking for the row that has 'Employee Number' etc."""
    for r in range(1, min(6, (sheet.max_row or 0) + 1)):
        values = [normalize_header(c.value) for c in sheet[r]]
        if any("employee number" in v or "first name" in v for v in values):
            return r
    return 1


def _detect_columns(sheet: Worksheet, header_row: int) -> dict[str, int]:
    headers = [normalize_header(c.value) for c in sheet[header_row]]
    cols: dict[str, int] = {}
    for field_name, patterns in COLUMN_KEYWORDS.items():
        for pat in patterns:
            for idx, h in enumerate(headers):
                if pat in h:
                    cols[field_name] = idx
                    break
            if field_name in cols:
                break
    return cols


def _row_to_employee(
    row: tuple[Any, ...],
    cols: dict[str, int],
    effective_date: date_type,
) -> ParsedEmployee | None:

    def get(name: str) -> Any:
        idx = cols.get(name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    emp_num_raw = get("emp_num")
    first = get("first_name")
    last = get("last_name")

    if not emp_num_raw and not first and not last:
        return None

    dob_iso = clean_date(get("dob"))
    # Prefer direct "age" column; fall back to DOB calculation
    age = clean_int(get("age")) if "age" in cols else _calc_age(dob_iso, effective_date)
    if age is None:
        age = _calc_age(dob_iso, effective_date)
    hours_pp = clean_float(get("hours_per_pay_period"))
    # Carlisle pay fortnightly → hours_per_week = hpp / 2
    hours_per_week = round(hours_pp / 2, 2) if hours_pp is not None else None

    return ParsedEmployee(
        emp_num=str(emp_num_raw or "").strip(),
        first_name=clean_str(first) or "",
        last_name=clean_str(last) or "",
        site=clean_str(get("site")) or "Unknown",
        preferred_name=clean_str(get("preferred_name")),
        email=clean_str(get("email")),
        dob=dob_iso,
        age=age,
        service_start_date=clean_date(get("service_start_date")),
        hire_date=clean_date(get("hire_date")),
        category=clean_str(get("category")),
        job_classification=clean_str(get("job_classification")),
        rate_type=clean_str(get("rate_type")),
        current_rate=clean_float(get("current_rate")),
        hours_per_pay_period=hours_pp,
        hours_per_week=hours_per_week,
        current_award=clean_str(get("current_award")),
    )


def _calc_age(dob_iso: str | None, effective_date: date_type) -> int | None:
    """Calculate age in years at the effective_date."""
    if not dob_iso:
        return None
    try:
        dob = datetime.fromisoformat(dob_iso).date()
    except (TypeError, ValueError):
        return None
    years = effective_date.year - dob.year
    if (effective_date.month, effective_date.day) < (dob.month, dob.day):
        years -= 1
    return years
