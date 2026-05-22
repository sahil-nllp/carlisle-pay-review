"""Excel wage-model parser.

Reads a Carlisle "Wage Report" workbook and returns structured records.

Column mapping (FY26 model):
  - emp_num        → Employee Number
  - site           → Carlisle Health (Site)
  - first_name     → First Name
  - last_name      → Last Name
  - fy26_award     → FY26 Award Agreement  (current award level for this cycle)
  - pp_level       → FY26 Level P&P & Notes (pay progression band)
  - current_rate   → NEW RATE              (the rate employees are currently on)
  - fy25_award     → FY25 Award Agreement  (prior year, for reference only)

NOTE: Change type, change input, proposed rate and letter type are NOT read from
Excel — they are calculated by the application during the review workflow.

Tolerates:
  - Column re-ordering (detects by header keyword)
  - Missing optional columns
  - Multiple sheet names — picks the wide one if there's no exact match
  - Blank rows
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


# ─────────────────────────────────────────────────────────────────────────────
#  Column detection
# ─────────────────────────────────────────────────────────────────────────────
COLUMN_KEYWORDS: dict[str, list[str]] = {
    "emp_num":      ["employee number", "emp no", "emp #", "emp number"],
    "site":         ["carlisle health(site)", "carlisle health (site)", "site"],
    "first_name":   ["first name"],
    "last_name":    ["last name", "surname"],
    "dob":          ["date of birth", "dob", "birth date"],
    "age":          ["^age$"],
    "category":     ["employee category"],
    "hours":        ["contracted hours", "hours/wk", "hours per week"],
    "email":        ["email"],
    "dept":         ["department"],
    # FY25 award kept for historical reference only — not used in calculations
    "fy25_award":   ["fy25 award", "fy2025 award", "fy25 award agreement"],
    # FY26 = current award for this cycle
    "fy26_award":   ["fy26 award agreement", "fy26 award", "fy2026 award", "new award agreement"],
    # Pay progression band description
    "pp_level":     ["fy26 level p&p", "fy26 level p&p & notes", "p&p description", "p&p level", "pay progression"],
    # NEW RATE is the employee's current rate entering this review
    "current_rate": ["new rate", "current rate"],
    # departed marker
    "change_col":   ["^change$", "change type", "review type"],
    "notes":        ["^notes$"],

    # ── Historical compliance snapshot (FY25→FY26) — display-only ────────────
    # Stored as-is from Excel; the app re-derives equivalent checks for FY26→FY27.
    "hist_award_level_changed": ["has award level changed"],
    "hist_rate_changed":        ["has rate changed"],
    "hist_above_award_rate":    ["same or above award rate", "above award rate"],
    "hist_above_pp_rate":       ["same or above pay progression rate", "above pay progression rate",
                                 "above pp rate"],
    "hist_above_pp_max":        ["same or above max pay progression", "above max pay progression",
                                 "above max pp"],
}

# Sheet candidates (case-insensitive substring match)
PREFERRED_SHEETS = ["wage report", "review", "master"]


# ─────────────────────────────────────────────────────────────────────────────
#  Result types
# ─────────────────────────────────────────────────────────────────────────────
@dataclass
class ParsedEmployee:
    """One row from the wage model."""

    emp_num: str
    first_name: str
    last_name: str
    site: str

    email: str | None = None
    dob: str | None = None             # ISO date string or None
    age: int | None = None
    category: str | None = None
    dept: str | None = None
    hours_per_week: float | None = None

    # Prior year award — reference only
    fy25_award: str | None = None

    # Current state for this FY26 review cycle (from Excel)
    current_rate: float | None = None  # NEW RATE column
    fy26_award: str | None = None      # FY26 Award Agreement
    pp_level: str | None = None        # FY26 Level P&P & Notes

    # Change/proposed fields are intentionally NOT read from Excel —
    # they are initialised with CPI defaults when the cycle is created.

    # Historical compliance snapshot — FY25→FY26 (from Excel, read-only)
    hist_award_level_changed: bool | None = None
    hist_rate_changed:        bool | None = None
    hist_above_award_rate:    bool | None = None
    hist_above_pp_rate:       bool | None = None
    hist_above_pp_max:        bool | None = None

    is_departed: bool = False


@dataclass
class ParseResult:
    sheet_name: str
    column_map: dict[str, int]              # field -> 0-based column index
    employees: list[ParsedEmployee] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


# ─────────────────────────────────────────────────────────────────────────────
#  Public API
# ─────────────────────────────────────────────────────────────────────────────
def parse_wage_model(source: bytes | Path | str) -> ParseResult:
    """Parse a wage model Excel file. Accepts bytes, path, or path-like."""
    if isinstance(source, (bytes, bytearray)):
        wb = openpyxl.load_workbook(BytesIO(source), data_only=True, read_only=False)
    else:
        wb = openpyxl.load_workbook(str(source), data_only=True, read_only=False)

    sheet = _pick_sheet(wb)
    if sheet is None:
        raise ValueError(
            "Could not find a usable sheet. Expected a sheet named 'Wage Report', "
            "'REVIEW' or 'MASTER', or a sheet with >20 columns."
        )

    header_row = _find_header_row(sheet)
    column_map = _detect_columns(sheet, header_row)

    required = {"emp_num", "first_name", "last_name", "site"}
    missing = required - column_map.keys()
    if missing:
        raise ValueError(
            f"Required columns not found: {sorted(missing)}. "
            f"Detected columns: {sorted(column_map.keys())}"
        )

    result = ParseResult(sheet_name=sheet.title, column_map=column_map)

    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        emp = _row_to_employee(row, column_map)
        if emp is None:
            continue
        result.employees.append(emp)

    if not result.employees:
        result.warnings.append("No data rows found in the sheet.")

    return result


# ─────────────────────────────────────────────────────────────────────────────
#  Internals
# ─────────────────────────────────────────────────────────────────────────────
def _pick_sheet(wb: openpyxl.Workbook) -> Worksheet | None:
    for name in wb.sheetnames:
        lname = name.lower()
        if any(p in lname for p in PREFERRED_SHEETS):
            return wb[name]
    candidates = [wb[n] for n in wb.sheetnames]
    candidates = [s for s in candidates if s.max_row and s.max_row > 1]
    if not candidates:
        return None
    return max(candidates, key=lambda s: s.max_column or 0)


def _find_header_row(sheet: Worksheet) -> int:
    for r in (1, 2, 3):
        cells = sheet[r] if r <= sheet.max_row else []
        values = [str(c.value or "").strip().lower() for c in cells]
        if any("first name" in v or "emp" in v or "site" in v for v in values):
            return r
    return 1


def _detect_columns(sheet: Worksheet, header_row: int) -> dict[str, int]:
    """Return {field: 0-based column index}."""
    header_cells = sheet[header_row]
    headers = [str(c.value or "").strip().lower() for c in header_cells]

    cols: dict[str, int] = {}
    for field_name, patterns in COLUMN_KEYWORDS.items():
        for pat in patterns:
            for idx, h in enumerate(headers):
                if not h:
                    continue
                if pat.startswith("^") and pat.endswith("$"):
                    if re.fullmatch(pat[1:-1], h):
                        cols[field_name] = idx
                        break
                elif pat.startswith("^"):
                    if re.match(pat[1:], h):
                        cols[field_name] = idx
                        break
                elif pat in h:
                    cols[field_name] = idx
                    break
            if field_name in cols:
                break
    return cols


def _row_to_employee(
    row: tuple[Any, ...], cols: dict[str, int]
) -> ParsedEmployee | None:

    def get(field_name: str) -> Any:
        idx = cols.get(field_name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    emp_num_raw = get("emp_num")
    first = get("first_name")
    last = get("last_name")

    if not emp_num_raw and not first and not last:
        return None

    # Departed detection — check the change column for "DEPART" marker
    change_marker = str(get("change_col") or "").upper()
    is_departed = "DEPART" in change_marker

    return ParsedEmployee(
        emp_num=str(emp_num_raw or "").strip(),
        first_name=_clean_str(first) or "",
        last_name=_clean_str(last) or "",
        site=_clean_str(get("site")) or "Unknown",
        email=_clean_str(get("email")),
        dob=_clean_date(get("dob")),
        age=_clean_int(get("age")),
        category=_clean_str(get("category")),
        dept=_clean_str(get("dept")),
        hours_per_week=_clean_float(get("hours")),
        fy25_award=_clean_str(get("fy25_award")),
        current_rate=_clean_float(get("current_rate")),  # NEW RATE column
        fy26_award=_clean_str(get("fy26_award")),        # FY26 Award Agreement
        pp_level=_clean_str(get("pp_level")),            # FY26 Level P&P & Notes
        # Historical compliance snapshot (FY25→FY26)
        hist_award_level_changed=_clean_bool(get("hist_award_level_changed")),
        hist_rate_changed=       _clean_bool(get("hist_rate_changed")),
        hist_above_award_rate=   _clean_bool(get("hist_above_award_rate")),
        hist_above_pp_rate=      _clean_bool(get("hist_above_pp_rate")),
        hist_above_pp_max=       _clean_bool(get("hist_above_pp_max")),
        is_departed=is_departed,
    )


# ─────────────────────────────────────────────────────────────────────────────
#  Value cleaners
# ─────────────────────────────────────────────────────────────────────────────
def _clean_bool(v: Any) -> bool | None:
    """Parse YES/NO/Y/N/True/False/1/0 cell values into bool, or None if blank."""
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("yes", "y", "true", "1"):
        return True
    if s in ("no", "n", "false", "0"):
        return False
    return None


def _clean_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _clean_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _clean_int(v: Any) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _clean_date(v: Any) -> str | None:
    if v is None or v == "":
        return None
    if hasattr(v, "isoformat"):
        return v.isoformat()[:10]
    s = str(v).strip()
    return s if s else None
