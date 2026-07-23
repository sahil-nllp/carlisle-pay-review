"""Downloads endpoints — Phase 5.

Routes:
  GET  /cycles/{cycle_id}/downloads                  — list all generated files for a cycle
  GET  /cycles/{cycle_id}/compliance-notes-report    — Excel of all marked-as-noted notes
  GET  /downloads/{file_id}                          — stream a file for download
"""
from __future__ import annotations

import io
from datetime import timezone
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.dependencies import get_current_user
from app.database import get_db
from app.models import ComplianceSuppression, Employee, GeneratedFile, User

router = APIRouter(tags=["downloads"])

FILE_TYPE_LABELS = {
    "letters_zip":     "Pay Letters (ZIP)",
    "ukg_upload":      "UKG Payroll Upload",
    "regional_excel":  "Regional Summary Excel",
    "mailmerge_a":     "Mail-Merge Template (Letter A)",
    "mailmerge_b":     "Mail-Merge Template (Letter B)",
    "mailmerge_c":     "Mail-Merge Template (Letter C)",
}

MIME_TYPES = {
    "letters_zip":    "application/zip",
    "ukg_upload":     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "regional_excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "mailmerge_a":    "application/vnd.ms-excel.sheet.macroEnabled.12",
    "mailmerge_b":    "application/vnd.ms-excel.sheet.macroEnabled.12",
    "mailmerge_c":    "application/vnd.ms-excel.sheet.macroEnabled.12",
}


@router.get("/cycles/{cycle_id}/downloads")
async def list_downloads(
    cycle_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> list[dict]:
    """Return all generated files for a cycle, newest first."""
    stmt = (
        select(GeneratedFile)
        .where(GeneratedFile.cycle_id == cycle_id)
        .order_by(GeneratedFile.created_at.desc())
    )
    result = await db.execute(stmt)
    files = result.scalars().all()

    return [
        {
            "id": f.id,
            "site": f.site,
            "file_type": f.file_type,
            "label": FILE_TYPE_LABELS.get(f.file_type, f.file_type),
            "filename": f.filename,
            "file_size": f.file_size,
            "created_at": f.created_at.isoformat() if f.created_at else None,
        }
        for f in files
    ]


@router.get("/cycles/{cycle_id}/compliance-notes-report")
async def compliance_notes_report(
    cycle_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> StreamingResponse:
    """Generate and stream an Excel listing all compliance notes for the cycle."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from zoneinfo import ZoneInfo
    AEST = ZoneInfo("Australia/Sydney")

    # Fetch all active suppressions for this cycle via employees
    emp_result = await db.execute(
        select(Employee).where(Employee.cycle_id == cycle_id)
    )
    employees = {e.id: e for e in emp_result.scalars().all()}

    if not employees:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "No employees found for this cycle")

    supp_result = await db.execute(
        select(ComplianceSuppression)
        .where(ComplianceSuppression.employee_id.in_(employees.keys()))
        .order_by(ComplianceSuppression.suppressed_at.asc())
    )
    suppressions = supp_result.scalars().all()

    # Collect user IDs to look up names
    user_ids = {s.suppressed_by_id for s in suppressions} | {s.undone_by_id for s in suppressions if s.undone_by_id}
    user_map: dict[int, str] = {}
    if user_ids:
        u_result = await db.execute(select(User).where(User.id.in_(user_ids)))
        for u in u_result.scalars():
            user_map[u.id] = u.name or u.email

    # Build Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compliance Notes"

    NAVY  = "1F3864"
    WHITE = "FFFFFF"
    GREY  = "F2F2F2"

    def fill(c: str) -> PatternFill:
        return PatternFill("solid", fgColor=c)

    headers = [
        ("Site",         20),
        ("Emp #",        10),
        ("Employee",     22),
        ("Check",        28),
        ("Status",       12),
        ("Note / Reason", 40),
        ("Noted by",     20),
        ("Noted at",     18),
        ("Undone by",    20),
        ("Undone at",    18),
    ]

    # Header row
    ws.row_dimensions[1].height = 30
    for col, (h, w) in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True, size=10, name="Calibri", color=WHITE)
        c.fill = fill(NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[chr(64 + col)].width = w

    def fmt_dt(dt) -> str:
        if not dt:
            return ""
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(AEST).strftime("%d/%m/%Y %H:%M")

    # Data rows
    for row_num, s in enumerate(suppressions, start=2):
        emp = employees.get(s.employee_id)
        bg = GREY if row_num % 2 == 0 else WHITE

        row_data = [
            emp.site if emp else "",
            emp.emp_num if emp else "",
            f"{emp.first_name} {emp.last_name}" if emp else "",
            s.check_label,
            "Active" if s.is_active else "Undone",
            s.reason or "",
            user_map.get(s.suppressed_by_id, str(s.suppressed_by_id)),
            fmt_dt(s.suppressed_at),
            user_map.get(s.undone_by_id, "") if s.undone_by_id else "",
            fmt_dt(s.undone_at) if s.undone_at else "",
        ]

        for col, val in enumerate(row_data, 1):
            c = ws.cell(row_num, col, val)
            c.font = Font(size=10, name="Calibri")
            c.fill = fill(bg)
            c.alignment = Alignment(vertical="center", wrap_text=col == 6)
        ws.row_dimensions[row_num].height = 16

    # Freeze header rows
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="compliance-notes-cycle-{cycle_id}.xlsx"'},
    )


@router.get("/downloads/{file_id}")
async def download_file(
    file_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> FileResponse:
    """Stream a generated file for download."""
    gf = await db.get(GeneratedFile, file_id)
    if not gf:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "File not found")

    path = Path(gf.file_path)
    if not path.exists():
        raise HTTPException(status.HTTP_404_NOT_FOUND, "File has been removed from storage")

    return FileResponse(
        path=str(path),
        filename=gf.filename,
        media_type=MIME_TYPES.get(gf.file_type, "application/octet-stream"),
    )
